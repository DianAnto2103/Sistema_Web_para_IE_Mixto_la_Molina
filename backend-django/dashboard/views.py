from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Count, Q
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from io import BytesIO
import json
from datetime import datetime
from api.models import Alumno, Curso, Calificacion, Taller, SolicitudTaller, Bimestre, Grado, Docente

# ========== LOGIN ==========
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            return render(request, 'dashboard/login.html', {'error': 'Credenciales inválidas'})
    return render(request, 'dashboard/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

# ========== DASHBOARD PRINCIPAL ==========
@login_required(login_url='login')
def home(request):
    total_alumnos = Alumno.objects.count()
    total_cursos = Curso.objects.count()
    total_talleres = Taller.objects.count()
    total_solicitudes = SolicitudTaller.objects.count()
    total_docentes = Docente.objects.count()

    # ========== PROMEDIO GENERAL ==========
    resultado_promedio = Calificacion.objects.aggregate(promedio=Avg('nota'))
    promedio_general = resultado_promedio['promedio']
    if promedio_general is None:
        promedio_general = 0
    else:
        promedio_general = float(round(promedio_general, 2))
    
    # ========== KPIs ESTRATÉGICOS ==========
    # Calcular promedios por alumno
    alumnos_con_promedio = []
    for alumno in Alumno.objects.all():
        promedio = Calificacion.objects.filter(alumno=alumno).aggregate(promedio=Avg('nota'))['promedio'] or 0
        alumnos_con_promedio.append({
            'alumno': alumno,
            'promedio': float(promedio)
        })
        
    
    # Porcentaje de alumnos destacados (≥16)
    destacados = sum(1 for a in alumnos_con_promedio if a['promedio'] >= 16)
    porcentaje_destacados = round((destacados / total_alumnos) * 100, 1) if total_alumnos > 0 else 0
    
    # Porcentaje de alumnos en riesgo (<11)
    en_riesgo = sum(1 for a in alumnos_con_promedio if a['promedio'] < 11)
    porcentaje_riesgo = round((en_riesgo / total_alumnos) * 100, 1) if total_alumnos > 0 else 0
    
    # Tasa de aprobados (≥11)
    aprobados = sum(1 for a in alumnos_con_promedio if a['promedio'] >= 11)
    tasa_aprobados = round((aprobados / total_alumnos) * 100, 1) if total_alumnos > 0 else 0

    
    # ========== COMPARATIVAS ==========
    # Promedios por grad
    promedios_grado = []
    
    orden_grados = [
        {'nombre': '1ro', 'nivel': 'Primaria'},
        {'nombre': '2do', 'nivel': 'Primaria'},
        {'nombre': '3ro', 'nivel': 'Primaria'},
        {'nombre': '4to', 'nivel': 'Primaria'},
        {'nombre': '5to', 'nivel': 'Primaria'},
        {'nombre': '6to', 'nivel': 'Primaria'},
        {'nombre': '1ro', 'nivel': 'Secundaria'},
        {'nombre': '2do', 'nivel': 'Secundaria'},
        {'nombre': '3ro', 'nivel': 'Secundaria'},
        {'nombre': '4to', 'nivel': 'Secundaria'},
        {'nombre': '5to', 'nivel': 'Secundaria'},
        {'nombre': '6to', 'nivel': 'Secundaria'},
    ]
    
    for grado_info in orden_grados:
        grado = Grado.objects.filter(nombre=grado_info['nombre'], nivel=grado_info['nivel']).first()
        if grado:
            promedio = Calificacion.objects.filter(alumno__grado=grado).aggregate(promedio=Avg('nota'))['promedio'] or 0
            texto_grado = f"{grado.nombre} de {grado.nivel}"
            promedios_grado.append({
                'grado': texto_grado,
                'promedio': float(round(promedio, 2))
            })
        else:
            texto_grado = f"{grado_info['nombre']} de {grado_info['nivel']}"
            promedios_grado.append({
                'grado': texto_grado,
                'promedio': 0
            })
        
        # Grado destacado y grado con menor rendimiento
        if promedios_grado:
            grado_destacado = max(promedios_grado, key=lambda x: x['promedio'])
            grado_riesgo = min(promedios_grado, key=lambda x: x['promedio'])
        else:
            grado_destacado = {'grado': 'N/A', 'promedio': 0}
            grado_riesgo = {'grado': 'N/A', 'promedio': 0}
    
    # Curso con menor rendimiento
    promedios_curso = []
    for curso in Curso.objects.all():
        promedio = Calificacion.objects.filter(curso=curso).aggregate(promedio=Avg('nota'))['promedio'] or 0
        promedios_curso.append({
            'curso': curso.nombre,
            'promedio': float(round(promedio, 2))
        })
    
    if promedios_curso:
        curso_riesgo = min(promedios_curso, key=lambda x: x['promedio'])
    else:
        curso_riesgo = {'curso': 'N/A', 'promedio': 0}
    
    # ========== TALLERES ==========
    solicitudes_aprobadas = SolicitudTaller.objects.filter(estado='aprobado').count()
    porcentaje_aprobadas = round((solicitudes_aprobadas / total_solicitudes) * 100, 1) if total_solicitudes > 0 else 0
    solicitudes_pendientes = SolicitudTaller.objects.filter(estado='pendiente').count()

    # ========== DISTRIBUCIÓN DE NOTAS ==========
    destacados = sum(1 for a in alumnos_con_promedio if a['promedio'] >= 16)
    aprobados = sum(1 for a in alumnos_con_promedio if 11 <= a['promedio'] < 16)
    riesgo = sum(1 for a in alumnos_con_promedio if a['promedio'] < 11)

    distribucion_notas = {
        'destacados': destacados,
        'aprobados': aprobados,
        'riesgo': riesgo,
        'porcentaje_destacados': round((destacados / total_alumnos) * 100, 1) if total_alumnos > 0 else 0,
        'porcentaje_aprobados': round((aprobados / total_alumnos) * 100, 1) if total_alumnos > 0 else 0,
        'porcentaje_riesgo': round((riesgo / total_alumnos) * 100, 1) if total_alumnos > 0 else 0,
    }
    
    # ========== EVOLUCIÓN ==========
    evolucion_data = []
    bimestres = Bimestre.objects.all().order_by('id')
    for bimestre in bimestres:
        promedio = Calificacion.objects.filter(bimestre=bimestre).aggregate(promedio=Avg('nota'))['promedio'] or 0
        evolucion_data.append({
            'bimestre': bimestre.nombre,
            'promedio': round(float(promedio), 2)
        })
    
    evolucion_json = json.dumps(evolucion_data)
    promedios_grado_json = json.dumps(promedios_grado)

    # ========== TENDENCIAS ==========
    tendencia = {}
    if len(evolucion_data) >= 2:
        ultimo = evolucion_data[-1]['promedio']
        anterior = evolucion_data[-2]['promedio']
        diferencia = ultimo - anterior
        tendencia = {
            'valor': round(diferencia, 2),
            'tipo': 'positiva' if diferencia > 0 else 'negativa' if diferencia < 0 else 'estable',
            'icono': '▲' if diferencia > 0 else '▼' if diferencia < 0 else '●',
            'color': 'success' if diferencia > 0 else 'danger' if diferencia < 0 else 'secondary'
        }
    else:
        tendencia = {'valor': 0, 'tipo': 'estable', 'icono': '●', 'color': 'secondary'}
    
    context = {
        # KPIs básicos
        'total_alumnos': total_alumnos,
        'total_cursos': total_cursos,
        'total_talleres': total_talleres,
        'total_solicitudes': total_solicitudes,
        'total_docentes': total_docentes,
        
        # KPIs estratégicos
        'porcentaje_destacados': porcentaje_destacados,
        'destacados_count': destacados,
        'porcentaje_riesgo': porcentaje_riesgo,
        'riesgo_count': en_riesgo,
        'tasa_aprobados': tasa_aprobados,
        'aprobados_count': aprobados,
        'distribucion_notas': distribucion_notas,
        'tendencia': tendencia,
        'promedio_general': promedio_general,
        
        # Comparativas
        'grado_destacado': grado_destacado,
        'grado_riesgo': grado_riesgo,
        'curso_riesgo': curso_riesgo,
        
        # Talleres
        'solicitudes_aprobadas': solicitudes_aprobadas,
        'porcentaje_aprobadas': porcentaje_aprobadas,
        'solicitudes_pendientes': solicitudes_pendientes,
        
        # Gráficos
        'promedios_grado': promedios_grado,
        'promedios_grado_json': promedios_grado_json,
        'evolucion_data': evolucion_data,
        'evolucion_json': evolucion_json,

    }

    return render(request, 'dashboard/dashboard_home.html', context)


# ========== EVOLUCIÓN DE NOTAS ==========
@login_required(login_url='login')
def evolucion(request):
    bimestres = Bimestre.objects.all().order_by('id')
    evolucion_data = []
    for bimestre in bimestres:
        promedio = Calificacion.objects.filter(bimestre=bimestre).aggregate(promedio=Avg('nota'))['promedio'] or 0
        evolucion_data.append({
            'id': bimestre.id,
            'bimestre': bimestre.nombre,
            'promedio': round(float(promedio), 2)  # Convertir Decimal a float
        })
    
    evolucion_json = json.dumps(evolucion_data)
    
    context = {
        'evolucion_data': evolucion_data,
        'evolucion_json': evolucion_json,
    }
    return render(request, 'dashboard/evolucion.html', context)

# ========== REPORTES ==========
@login_required(login_url='login')
def reportes(request):
    grados = Alumno.objects.values('grado__id', 'grado__nombre').distinct()
    context = {
        'grados': grados,
    }
    return render(request, 'dashboard/reportes.html', context)

# ========== EXPORTAR REPORTES ==========
@login_required(login_url='login')
def exportar_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_mixtotrack.pdf"'
    
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "MixtoTrack - Reporte Académico")
    
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 80, f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Tabla de alumnos
    y = height - 120
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "ID")
    p.drawString(100, y, "Apellidos")
    p.drawString(250, y, "Nombres")
    p.drawString(400, y, "Promedio")
    
    y -= 20
    p.setFont("Helvetica", 10)
    
    alumnos = Alumno.objects.all()
    for alumno in alumnos:
        promedio = Calificacion.objects.filter(alumno=alumno).aggregate(promedio=Avg('nota'))['promedio'] or 0
        p.drawString(50, y, str(alumno.id))
        p.drawString(100, y, alumno.apellidos[:30])
        p.drawString(250, y, alumno.nombres[:30])
        p.drawString(400, y, f"{float(promedio):.2f}")
        y -= 20
        if y < 50:
            p.showPage()
            y = height - 50
    
    p.showPage()
    p.save()
    return response

@login_required(login_url='login')
def exportar_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_mixtotrack.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Académico"
    
    # Encabezados
    headers = ["ID", "Apellidos", "Nombres", "Grado", "Sección", "Promedio General"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos
    row = 2
    alumnos = Alumno.objects.all()
    for alumno in alumnos:
        promedio = Calificacion.objects.filter(alumno=alumno).aggregate(promedio=Avg('nota'))['promedio'] or 0
        ws.cell(row=row, column=1, value=alumno.id)
        ws.cell(row=row, column=2, value=alumno.apellidos)
        ws.cell(row=row, column=3, value=alumno.nombres)
        ws.cell(row=row, column=4, value=alumno.grado.nombre if alumno.grado else "")
        ws.cell(row=row, column=5, value=alumno.seccion or "")
        ws.cell(row=row, column=6, value=round(float(promedio), 2))
        row += 1
    
    wb.save(response)
    return response

@login_required(login_url='login')
def exportar_txt(request):
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename="reporte_mixtotrack.txt"'
    
    content = "MIXTOTRACK - REPORTE ACADEMICO\n"
    content += "=" * 50 + "\n\n"
    
    alumnos = Alumno.objects.all()
    for alumno in alumnos:
        promedio = Calificacion.objects.filter(alumno=alumno).aggregate(promedio=Avg('nota'))['promedio'] or 0
        content += f"ID: {alumno.id}\n"
        content += f"Apellidos: {alumno.apellidos}\n"
        content += f"Nombres: {alumno.nombres}\n"
        content += f"Promedio: {float(promedio):.2f}\n"
        content += "-" * 30 + "\n"
    
    response.write(content)
    return response


@login_required(login_url='login')
def estudiantes_list(request):
    from django.db.models import Avg
    from api.models import Alumno, Calificacion
    
    alumnos = Alumno.objects.all().select_related('grado')
    
    # Calcular promedio de CADA alumno (esto va dentro del bucle)
    for alumno in alumnos:
        promedio = Calificacion.objects.filter(alumno=alumno).aggregate(promedio=Avg('nota'))['promedio'] or 0
        alumno.promedio_general = float(round(promedio, 2))
    
    context = {
        'alumnos': alumnos,
    }
    return render(request, 'dashboard/estudiantes_list.html', context)


@login_required(login_url='login')
def estudiante_create(request):
    if request.method == 'POST':
        alumno = Alumno(
            nombres=request.POST.get('nombres'),
            apellidos=request.POST.get('apellidos'),
            correo=request.POST.get('correo'),
            grado_id=request.POST.get('grado') or None,
            seccion=request.POST.get('seccion'),
        )
        alumno.save()
        return redirect('estudiantes_list')
    
    grados = Grado.objects.all()
    return render(request, 'dashboard/estudiante_form.html', {'grados': grados})

@login_required(login_url='login')
def rendimiento_estudiante(request):
    alumno_id = request.GET.get('alumno_id')
    bimestre_id = request.GET.get('bimestre_id')
    
    if not alumno_id:
        return redirect('estudiantes_list')
    
    alumno = Alumno.objects.get(id=alumno_id)
    bimestres = Bimestre.objects.all().order_by('id')
    
    # Base de calificaciones
    calificaciones_base = Calificacion.objects.filter(alumno=alumno).select_related('curso', 'bimestre')
    
    # Filtrar por bimestre si se seleccionó uno
    if bimestre_id and bimestre_id.isdigit():
        calificaciones = calificaciones_base.filter(bimestre_id=bimestre_id)
        bimestre_seleccionado = int(bimestre_id)
        bimestre_seleccionado_nombre = Bimestre.objects.get(id=bimestre_id).nombre
        promedio_bimestre = calificaciones.aggregate(promedio=Avg('nota'))['promedio'] or 0
        promedio_bimestre = round(float(promedio_bimestre), 2)
    else:
        calificaciones = calificaciones_base
        bimestre_seleccionado = None
        bimestre_seleccionado_nombre = None
        promedio_bimestre = None
    
    # Estadísticas del bimestre seleccionado
    aprobados = 0
    regulares = 0
    desaprobados = 0
    
    for cal in calificaciones:
        if cal.nota >= 14:
            aprobados += 1
        elif cal.nota >= 11:
            regulares += 1
        else:
            desaprobados += 1
    
    # Promedio general (anual)
    promedio_general = calificaciones_base.aggregate(promedio=Avg('nota'))['promedio'] or 0
    promedio_general = round(float(promedio_general), 2)
    
    # Datos para el gráfico de evolución (JSON)
    evolucion = []
    for bim in bimestres:
        promedio = calificaciones_base.filter(bimestre=bim).aggregate(promedio=Avg('nota'))['promedio'] or 0
        evolucion.append({
            'bimestre': bim.nombre,
            'promedio': round(float(promedio), 2)
        })
    evolucion_json = json.dumps(evolucion)
    
    context = {
        'alumno': alumno,
        'calificaciones': calificaciones,
        'bimestres': bimestres,
        'bimestre_seleccionado': bimestre_seleccionado,
        'bimestre_seleccionado_nombre': bimestre_seleccionado_nombre,
        'promedio_bimestre': promedio_bimestre,
        'promedio_general': promedio_general,
        'total_cursos': calificaciones.count(),
        'aprobados': aprobados,
        'regulares': regulares,
        'desaprobados': desaprobados,
        'evolucion_json': evolucion_json,
    }
    return render(request, 'dashboard/rendimiento_estudiante.html', context)


@login_required(login_url='login')
def exportar_pdf_estudiante(request):
    alumno_id = request.GET.get('alumno_id')
    if not alumno_id:
        return redirect('rendimiento_estudiante')
    
    alumno = Alumno.objects.get(id=alumno_id)
    calificaciones = Calificacion.objects.filter(alumno=alumno).select_related('curso', 'bimestre')
    
    if not calificaciones.exists():
        return HttpResponse("No hay calificaciones para este estudiante", status=404)
    
    promedio_general = calificaciones.aggregate(promedio=Avg('nota'))['promedio'] or 0
    
    # Agrupar notas por bimestre
    bimestres = Bimestre.objects.all().order_by('id')
    notas_por_bimestre = {}
    for bim in bimestres:
        notas_por_bimestre[bim.id] = {
            'nombre': bim.nombre,
            'notas': [],
            'promedio': 0
        }
    
    for cal in calificaciones:
        if cal.bimestre and cal.bimestre.id in notas_por_bimestre:
            notas_por_bimestre[cal.bimestre.id]['notas'].append({
                'curso': cal.curso.nombre,
                'nota': float(cal.nota)
            })
    
    # Calcular promedios por bimestre
    for bim_id in notas_por_bimestre:
        notas = notas_por_bimestre[bim_id]['notas']
        if notas:
            suma = sum(n['nota'] for n in notas)
            notas_por_bimestre[bim_id]['promedio'] = round(suma / len(notas), 2)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="boletin_{alumno.apellidos}_{alumno.nombres}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter, 
                            topMargin=0.7*inch, bottomMargin=0.7*inch,
                            leftMargin=0.7*inch, rightMargin=0.7*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Estilos personalizados
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#2E7D32'), alignment=1, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=11, textColor=colors.grey, alignment=1, spaceAfter=20)
    section_style = ParagraphStyle('Section', parent=styles['Heading3'], fontSize=13, textColor=colors.black, alignment=0, spaceAfter=8, spaceBefore=15)
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, leading=14)
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)
    
    # Encabezado
    story.append(Paragraph("MIXTOTRACK", title_style))
    story.append(Paragraph("Sistema de Gestión Académica - Libreta de Notas", subtitle_style))
    story.append(Spacer(1, 5))
    
    # Línea separadora
    story.append(Paragraph("-" * 80, styles['Normal']))
    story.append(Spacer(1, 10))
    
    # Datos del estudiante
    story.append(Paragraph(f"<b>ESTUDIANTE:</b> {alumno.apellidos}, {alumno.nombres}", info_style))
    grado_text = f"{alumno.grado.nombre if alumno.grado else '-'} {alumno.grado.nivel if alumno.grado and alumno.grado.nivel else ''}".strip()
    story.append(Paragraph(f"<b>GRADO Y SECCIÓN:</b> {grado_text} - {alumno.seccion or '-'}", info_style))
    story.append(Paragraph(f"<b>FECHA DE EMISIÓN:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", info_style))
    story.append(Spacer(1, 15))
    
    # Tabla de notas por bimestre
    for bim_id, bim_data in notas_por_bimestre.items():
        if bim_data['notas']:
            story.append(Paragraph(f"▸ {bim_data['nombre']}", section_style))
            
            table_data = [['CURSO', 'NOTA']]
            for nota in bim_data['notas']:
                nota_val = nota['nota']
                table_data.append([nota['curso'], f"{nota_val:.2f}"])
            
            # Fila de promedio
            table_data.append(['', ''])
            table_data.append(['PROMEDIO DEL BIMESTRE', f"{bim_data['promedio']:.2f}"])
            
            table = Table(table_data, colWidths=[4*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            story.append(table)
            story.append(Spacer(1, 15))
    
    # Resumen General
    story.append(Paragraph("RESUMEN GENERAL", section_style))
    
    total_cursos = calificaciones.count()
    aprobados = sum(1 for c in calificaciones if c.nota >= 14)
    regulares = sum(1 for c in calificaciones if 11 <= c.nota < 14)
    desaprobados = sum(1 for c in calificaciones if c.nota < 11)
    
    resumen_data = [
        ["Total de Cursos Evaluados:", str(total_cursos)],
        ["Cursos Aprobados (≥14):", f"{aprobados} ({round(aprobados/total_cursos*100,1) if total_cursos > 0 else 0}%)"],
        ["Cursos Regulares (11-13):", f"{regulares} ({round(regulares/total_cursos*100,1) if total_cursos > 0 else 0}%)"],
        ["Cursos Desaprobados (≤10):", f"{desaprobados} ({round(desaprobados/total_cursos*100,1) if total_cursos > 0 else 0}%)"],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
    resumen_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
    ]))
    story.append(resumen_table)
    story.append(Spacer(1, 10))
    
    # Promedio General destacado
    story.append(Spacer(1, 5))
    promedio_style = ParagraphStyle('Promedio', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#2E7D32'), alignment=1, spaceAfter=10)
    story.append(Paragraph(f"<b>PROMEDIO GENERAL: {float(promedio_general):.2f}</b>", promedio_style))
    story.append(Spacer(1, 15))
    
    # Pie de página
    story.append(Paragraph("-" * 80, styles['Normal']))
    story.append(Paragraph("Documento generado electrónicamente por MixtoTrack", footer_style))
    story.append(Paragraph(f"© {datetime.now().year} - I.E. Mixto La Molina", footer_style))
    
    doc.build(story)
    return response


@login_required(login_url='login')
def ranking(request):
    # Obtener todos los alumnos con sus promedios
    ranking_list = []
    for alumno in Alumno.objects.all().select_related('grado'):
        promedio = Calificacion.objects.filter(alumno=alumno).aggregate(promedio=Avg('nota'))['promedio'] or 0
        ranking_list.append({
            'id': alumno.id,
            'nombres': alumno.nombres,
            'apellidos': alumno.apellidos,
            'grado': alumno.grado,
            'seccion': alumno.seccion,
            'promedio_general': round(float(promedio), 2)
        })
    
    # Ordenar de mayor a menor promedio
    ranking_list.sort(key=lambda x: x['promedio_general'], reverse=True)
    
    # Top 10
    top_10 = ranking_list[:10]
    
    # Estadísticas adicionales para el dashboard de ranking
    promedio_general = 0
    if ranking_list:
        suma = sum(item['promedio_general'] for item in ranking_list)
        promedio_general = round(suma / len(ranking_list), 2)
    
    # Mejor y peor promedio
    mejor = ranking_list[0] if ranking_list else None
    peor = ranking_list[-1] if ranking_list else None
    
    # Distribución por nivel de rendimiento
    destacados = sum(1 for item in ranking_list if item['promedio_general'] >= 16)
    regulares = sum(1 for item in ranking_list if 11 <= item['promedio_general'] < 16)
    riesgo = sum(1 for item in ranking_list if item['promedio_general'] < 11)
    
    context = {
        'ranking': top_10,
        'total_alumnos': len(ranking_list),
        'promedio_general': promedio_general,
        'mejor': mejor,
        'peor': peor,
        'destacados': destacados,
        'regulares': regulares,
        'riesgo': riesgo,
    }
    return render(request, 'dashboard/ranking.html', context)